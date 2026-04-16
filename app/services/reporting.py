from collections import defaultdict
from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

from app.models import ResultSummary, Student


def parse_university_excel(file_storage, exam_name, semester, college_id=None):
    workbook = load_workbook(file_storage.stream)
    sheet = workbook.active
    records = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        register_no, total_marks, percentage, result_status = row[:4]
        if not register_no:
            continue
        student_query = Student.query.filter_by(register_no=str(register_no).strip())
        if college_id:
            student_query = student_query.filter_by(college_id=college_id)
        student = student_query.first()
        if not student:
            continue
        records.append(
            ResultSummary(
                college_id=student.college_id,
                student_id=student.id,
                result_type="University",
                exam_name=exam_name,
                semester=str(semester),
                total_marks=float(total_marks or 0),
                percentage=float(percentage or 0),
                result_status=str(result_status or "Pending"),
            )
        )
    return records


def build_excel_report(students, attendance_records, result_summaries):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "College Report"
    ws.append(["Student Name", "Register No", "Department", "Year", "Attendance %"])

    attendance_map = defaultdict(list)
    for record in attendance_records:
        attendance_map[record.student_id].append(record.status)

    for student in students:
        records = attendance_map.get(student.id, [])
        percent = 0
        if records:
            percent = round((records.count("Present") / len(records)) * 100, 2)
        ws.append([
            student.name,
            student.register_no,
            student.department.name,
            student.year,
            percent,
        ])

    ws2 = workbook.create_sheet("Results")
    ws2.append(["Student", "Exam", "Semester", "Type", "Total", "Percentage", "Status"])
    for summary in result_summaries:
        ws2.append([
            summary.student.name,
            summary.exam_name,
            summary.semester,
            summary.result_type,
            summary.total_marks,
            summary.percentage,
            summary.result_status,
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_pdf_report(students, attendance_records, result_summaries, college_name):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f"{college_name} - College Report", styles["Title"]), Spacer(1, 12)]

    attendance_map = defaultdict(list)
    for record in attendance_records:
        attendance_map[record.student_id].append(record.status)

    student_rows = [["Student", "Register No", "Department", "Attendance %"]]
    for student in students:
        records = attendance_map.get(student.id, [])
        percent = 0
        if records:
            percent = round((records.count("Present") / len(records)) * 100, 2)
        student_rows.append([student.name, student.register_no, student.department.name, f"{percent}%"])

    student_table = Table(student_rows, repeatRows=1)
    student_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3FAF")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ]
        )
    )
    elements.extend([student_table, Spacer(1, 16)])

    result_rows = [["Student", "Exam", "Type", "Percentage", "Status"]]
    for summary in result_summaries:
        result_rows.append([
            summary.student.name,
            summary.exam_name,
            summary.result_type,
            f"{summary.percentage}%",
            summary.result_status,
        ])

    result_table = Table(result_rows, repeatRows=1)
    result_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ]
        )
    )
    elements.append(result_table)
    doc.build(elements)
    output.seek(0)
    return output


def build_student_report_pdf(student, college_name):
    attendance_records = student.attendance_records
    result_summaries = student.summaries
    return build_pdf_report([student], attendance_records, result_summaries, college_name)
